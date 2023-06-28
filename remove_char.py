


    

from openpyxl import load_workbook

# Load the existing workbook
workbook = load_workbook('D:/pk.xlsx')
print(workbook.sheetnames)

# Select the worksheet you want to read and modify
worksheet = workbook['Issue Navigator']

# Get the column containing the dates and the column where you want to append the separated dates
date_column = worksheet['K']
new_column = worksheet['L']

# Iterate over the cells in the date column
for cell in date_column:
    # Access the cell value
    cell_value = cell.value
    
    # Split the date and time (assuming they are separated by a space)
    split_values = cell_value.split(' ')
    
    # Get the date part
    date_part = split_values[0]
    
    # Append the date part to the new column
    new_column.append([date_part])
    for index, value in enumerate(split_values):
        new_column.cell(row=index+1, column=1).value = value[0]

# Save the modified workbook
workbook.save('your_file.xlsx')

